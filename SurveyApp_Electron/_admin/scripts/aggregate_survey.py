#!/usr/bin/env python3
"""
アンケート集計処理スクリプト

機能:
1. config/内のJSONから Excelテンプレート生成
2. data/_csv/内のCSVを検出して統合
3. 重複IDは黄色ハイライト

使い方:
- 集計処理.bat をダブルクリック
"""

import json
import csv
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# === 定数 ===
CIRCLE_NUMBERS = [
    '⓪', '①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩',
    '⑪', '⑫', '⑬', '⑭', '⑮', '⑯', '⑰', '⑱', '⑲', '⑳',
    '㉑', '㉒', '㉓', '㉔', '㉕', '㉖', '㉗', '㉘', '㉙', '㉚',
    '㉛', '㉜', '㉝', '㉞', '㉟', '㊱', '㊲', '㊳', '㊴', '㊵',
    '㊶', '㊷', '㊸', '㊹', '㊺', '㊻', '㊼', '㊽', '㊾', '㊿'
]

BORDER_MEDIUM = Side(style='medium')
BORDER_THIN = Side(style='thin')
BORDER_DASHED = Side(style='dashed')

FILL_DUPLICATE = PatternFill('solid', fgColor='FFFF00')
FILL_ODD = PatternFill('solid', fgColor='FFFFFF')
FILL_EVEN = PatternFill('solid', fgColor='F2F2F2')


def to_circle_number(num):
    if 0 <= num < len(CIRCLE_NUMBERS):
        return CIRCLE_NUMBERS[num]
    return f'({num})'


# ============================================
# JSON解析
# ============================================
def parse_json_config(json_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    questions_config = {}
    questions_metadata = {}
    field_order = ['ID']
    other_fields = []
    
    def process_question(q, section_title):
        q_id = q['id']
        q_type = q.get('type', 'radio')
        options = q.get('options', [])
        csv_meta = q.get('csvMeta', {})
        section = csv_meta.get('section', '')
        title = csv_meta.get('title', q_id)
        
        if q_type == 'date_wareki':
            for suffix, suffix_title in [('_era', '_元号'), ('_year', '_年'), ('_month', '_月'), ('_day', '_日')]:
                fid = f"{q_id}{suffix}"
                field_order.append(fid)
                questions_metadata[fid] = {'section': section, 'title': f"{title}{suffix_title}"}
        elif q_type == 'year_month':
            for suffix, suffix_title in [('_year', '_年'), ('_month', '_月')]:
                fid = f"{q_id}{suffix}"
                field_order.append(fid)
                questions_metadata[fid] = {'section': section, 'title': f"{title}{suffix_title}"}
        elif q_type == 'number_pair':
            for field in q.get('fields', []):
                fid = field['id']
                field_order.append(fid)
                f_meta = field.get('csvMeta', {})
                questions_metadata[fid] = {
                    'section': f_meta.get('section', section),
                    'title': f_meta.get('title', fid)
                }
        elif q_type in ['radio', 'checkbox'] and options:
            values = [opt['value'] for opt in options]
            questions_config[q_id] = {'values': values}
            field_order.append(q_id)
            questions_metadata[q_id] = {'section': section, 'title': title}
            for opt in options:
                if opt.get('hasOther'):
                    other_id = f"{q_id}_other_{opt['value']}"
                    field_order.append(other_id)
                    other_fields.append(other_id)
        elif q_type == 'table':
            columns = q.get('columns', [])
            for row in q.get('rows', []):
                row_id = f"{q_id}_{row['id']}"
                values = [col['value'] for col in columns]
                questions_config[row_id] = {'values': values}
                field_order.append(row_id)
                row_meta = row.get('csvMeta', {})
                questions_metadata[row_id] = {
                    'section': row_meta.get('section', section),
                    'title': row_meta.get('title', row['label'])
                }
        elif q_type == 'scale':
            field_order.append(q_id)
            questions_metadata[q_id] = {'section': section, 'title': title}
        else:
            field_order.append(q_id)
            questions_metadata[q_id] = {'section': section, 'title': title}
        
        for sub in q.get('subQuestions', []):
            process_question(sub, section_title)
    
    for section in config.get('sections', []):
        section_title = section.get('title', '')
        for q in section.get('questions', []):
            process_question(q, section_title)
    
    field_order.append('入力日時')
    field_order.append('入力者')
    
    return {
        'survey_name': config.get('surveyName', 'survey'),
        'storage_key': config.get('storageKey', 'survey'),
        'questions_config': questions_config,
        'questions_metadata': questions_metadata,
        'field_order': field_order,
        'other_fields': other_fields
    }


# ============================================
# ヘッダー構築
# ============================================
def build_headers(parsed):
    headers = []
    row1 = []
    row2 = []
    row3 = []
    
    questions_config = parsed['questions_config']
    questions_metadata = parsed['questions_metadata']
    field_order = parsed['field_order']
    fields_set = set(field_order)
    
    last_section = ''
    col_idx = 1
    section_ranges = []
    current_section_start = None
    current_section_name = None
    
    for field in field_order:
        # otherフィールドはスキップ（選択肢展開時に挿入するため）
        if '_other_' in field:
            continue
        
        meta = questions_metadata.get(field, {})
        config = questions_config.get(field)
        section = meta.get('section', '')
        title = meta.get('title', field)
        
        if section and section != last_section:
            if current_section_name:
                section_ranges.append((current_section_name, current_section_start, col_idx - 1))
            current_section_name = section
            current_section_start = col_idx
            last_section = section
        
        if config and config.get('values'):
            values = config['values']
            for i, v in enumerate(values):
                headers.append({'field': field, 'value': v, 'col': col_idx})
                row1.append('')
                row2.append(title if i == 0 else '')
                row3.append(to_circle_number(v))
                col_idx += 1
                
                # この選択肢にotherがあれば直後に追加
                other_id = f"{field}_other_{v}"
                if other_id in fields_set:
                    headers.append({'field': other_id, 'col': col_idx})
                    row1.append('')
                    row2.append('')
                    row3.append('その他')
                    col_idx += 1
        else:
            headers.append({'field': field, 'col': col_idx})
            row1.append('')
            if field == 'ID':
                row2.append('NO')
            elif field == '入力日時':
                row2.append('入力日時')
            elif field == '入力者':
                row2.append('入力者')
            else:
                row2.append(title)
            row3.append('')
            col_idx += 1
    
    if current_section_name:
        section_ranges.append((current_section_name, current_section_start, col_idx - 1))
    
    return {
        'headers': headers,
        'row1': row1,
        'row2': row2,
        'row3': row3,
        'section_ranges': section_ranges,
        'total_cols': col_idx - 1
    }


# ============================================
# テンプレート生成
# ============================================
def create_template(parsed, output_path):
    header_info = build_headers(parsed)
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'データ'
    
    row1 = header_info['row1']
    row2 = header_info['row2']
    row3 = header_info['row3']
    section_ranges = header_info['section_ranges']
    total_cols = header_info['total_cols']
    headers = header_info['headers']
    
    font_normal = Font(name='游ゴシック', size=10)
    font_header = Font(name='游ゴシック', size=10, bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    question_starts = {}
    for h in headers:
        field = h['field']
        col = h['col']
        # otherフィールドは親フィールドの一部として扱う（設問開始としてマークしない）
        if '_other_' in field:
            continue
        if field not in question_starts:
            question_starts[field] = col
    
    # 1行目: 大問
    ws.cell(row=1, column=1).value = 'NO'
    ws.cell(row=1, column=1).font = font_header
    ws.cell(row=1, column=1).alignment = align_center
    ws.cell(row=1, column=1).border = Border(top=BORDER_MEDIUM, bottom=BORDER_MEDIUM, 
                                              left=BORDER_MEDIUM, right=BORDER_MEDIUM)
    
    for section_name, start_col, end_col in section_ranges:
        if start_col == 1:
            continue
        cell = ws.cell(row=1, column=start_col)
        cell.value = section_name
        cell.font = font_header
        cell.alignment = align_left
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        for col in range(start_col, end_col + 1):
            c = ws.cell(row=1, column=col)
            left = BORDER_MEDIUM if col == start_col else None
            right = BORDER_MEDIUM if col == end_col else None
            c.border = Border(top=BORDER_MEDIUM, bottom=BORDER_MEDIUM, left=left, right=right)
    
    for col in range(2, total_cols + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value is None:
            from openpyxl.cell.cell import MergedCell
            if not isinstance(cell, MergedCell):
                in_section = any(start <= col <= end for _, start, end in section_ranges)
                if not in_section:
                    cell.font = font_header
                    cell.border = Border(top=BORDER_MEDIUM, bottom=BORDER_MEDIUM, 
                                         left=BORDER_MEDIUM, right=BORDER_MEDIUM)
    
    ws.merge_cells('A1:A3')
    
    # 2行目: 小問
    # field_ranges: 各設問の開始列〜終了列（otherフィールドは親フィールドの一部として扱う）
    field_ranges = {}
    for h in headers:
        field = h['field']
        col = h['col']
        
        # otherフィールドは親フィールドの範囲に含める
        if '_other_' in field:
            # 親フィールド名を抽出（例: Q1_other_4 → Q1）
            parent_field = field.rsplit('_other_', 1)[0]
            if parent_field in field_ranges:
                field_ranges[parent_field][1] = max(field_ranges[parent_field][1], col)
            continue
        
        if field not in field_ranges:
            field_ranges[field] = [col, col]
        else:
            field_ranges[field][1] = col
    
    for col, val in enumerate(row2, start=1):
        if col == 1:
            continue
        cell = ws.cell(row=2, column=col)
        cell.font = font_normal
        is_question_start = col in question_starts.values()
        left = BORDER_MEDIUM if is_question_start else None
        cell.border = Border(top=BORDER_MEDIUM, bottom=BORDER_THIN, left=left)
    
    for field, (start_col, end_col) in field_ranges.items():
        if start_col == 1:
            continue
        meta = parsed['questions_metadata'].get(field, {})
        title = meta.get('title', field)
        if field == 'ID':
            title = 'NO'
        elif field == '入力日時':
            title = '入力日時'
        elif field == '入力者':
            title = '入力者'
        
        cell = ws.cell(row=2, column=start_col)
        cell.value = title
        cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)
        if end_col > start_col:
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
    
    # 3行目: 丸数字
    for col, val in enumerate(row3, start=1):
        if col == 1:
            continue
        cell = ws.cell(row=3, column=col)
        cell.value = val
        cell.font = font_normal
        cell.alignment = align_center
        is_question_start = col in question_starts.values()
        if is_question_start:
            left = BORDER_MEDIUM
        elif val:
            left = BORDER_DASHED
        else:
            left = BORDER_MEDIUM
        cell.border = Border(top=BORDER_THIN, bottom=BORDER_MEDIUM, left=left)
    
    for row in range(1, 4):
        cell = ws.cell(row=row, column=total_cols)
        old_border = cell.border
        cell.border = Border(top=old_border.top, bottom=old_border.bottom, 
                             left=old_border.left, right=BORDER_MEDIUM)
    
    # データ行（空）
    for row_offset in range(2):
        dest_row = 4 + row_offset
        fill = FILL_ODD if row_offset % 2 == 0 else FILL_EVEN
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=dest_row, column=col)
            cell.font = font_normal
            cell.alignment = align_center
            cell.fill = fill
            is_question_start = col in question_starts.values()
            left = BORDER_MEDIUM if is_question_start else BORDER_DASHED
            cell.border = Border(bottom=BORDER_THIN, left=left)
        ws.cell(row=dest_row, column=total_cols).border = Border(
            bottom=BORDER_THIN, left=BORDER_DASHED, right=BORDER_MEDIUM
        )
    
    # 列幅
    for col in range(1, total_cols + 1):
        col_letter = get_column_letter(col)
        if col == 1:
            ws.column_dimensions[col_letter].width = 6
        elif row3[col - 1]:
            ws.column_dimensions[col_letter].width = 3.5
        else:
            ws.column_dimensions[col_letter].width = 8
    
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.freeze_panes = 'A4'
    
    wb.save(output_path)
    return total_cols


# ============================================
# CSV読み込み
# ============================================
def read_csv_data(csv_path):
    encodings = ['utf-8-sig', 'utf-8', 'cp932', 'shift_jis']
    for encoding in encodings:
        try:
            with open(csv_path, 'r', encoding=encoding, newline='') as f:
                reader = csv.reader(f)
                rows = list(reader)
                if len(rows) <= 4:
                    return []
                return rows[4:]
        except UnicodeDecodeError:
            continue
    return []


# ============================================
# CSV統合
# ============================================
def merge_csv_to_template(template_path, csv_paths, output_path):
    wb = load_workbook(template_path)
    ws = wb.active
    total_cols = ws.max_column
    
    # テンプレートの2行目から設問開始位置を特定
    # （値があるセル = 設問タイトルの開始位置）
    question_starts = set()
    question_starts.add(1)  # ID列は常に設問開始
    for col in range(2, total_cols + 1):
        cell_value = ws.cell(row=2, column=col).value
        if cell_value and str(cell_value).strip():
            question_starts.add(col)
    
    existing_data_end = 3
    for row in range(4, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if cell.value is not None and str(cell.value).strip():
            existing_data_end = row
    
    start_row = existing_data_end + 1 if existing_data_end > 3 else 4
    
    existing_ids = set()
    for row in range(4, existing_data_end + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val:
            existing_ids.add(str(cell_val).strip())
    
    all_data = []
    for csv_path in csv_paths:
        data = read_csv_data(csv_path)
        all_data.extend(data)
    
    if not all_data:
        wb.close()
        return 0, 0
    
    all_ids = {}
    duplicate_rows = set()
    for i, row_data in enumerate(all_data):
        if row_data:
            row_id = str(row_data[0]).strip() if row_data[0] else ''
            if row_id:
                if row_id not in all_ids:
                    all_ids[row_id] = []
                all_ids[row_id].append(start_row + i)
                if row_id in existing_ids or len(all_ids[row_id]) > 1:
                    duplicate_rows.update(all_ids[row_id])
    
    font_normal = Font(name='游ゴシック', size=10)
    align_center = Alignment(horizontal='center', vertical='center')
    
    for i, row_data in enumerate(all_data):
        row_num = start_row + i
        is_odd = (row_num - 4) % 2 == 0
        is_duplicate = row_num in duplicate_rows
        fill = FILL_DUPLICATE if is_duplicate else (FILL_ODD if is_odd else FILL_EVEN)
        
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            
            # 値を設定
            if col <= len(row_data):
                cell.value = row_data[col - 1] if row_data[col - 1] else ''
            else:
                cell.value = ''
            
            cell.font = font_normal
            cell.alignment = align_center
            cell.fill = fill
            
            # 罫線設定（設問開始位置は太線、それ以外は破線）
            left_border = BORDER_MEDIUM if col in question_starts else BORDER_DASHED
            right_border = BORDER_MEDIUM if col == total_cols else None
            cell.border = Border(bottom=BORDER_THIN, left=left_border, right=right_border)
    
    wb.save(output_path)
    wb.close()
    
    return len(all_data), len(duplicate_rows)


# ============================================
# メイン処理
# ============================================
def main():
    print("=" * 50)
    print("  アンケート集計処理")
    print("=" * 50)
    print()
    
    # スクリプトの場所から基準パスを決定
    script_dir = Path(__file__).parent  # _admin/scripts/
    admin_dir = script_dir.parent        # _admin/
    base_dir = admin_dir.parent          # SurveyApp_template/
    config_dir = base_dir / 'config'
    csv_dir = base_dir / 'data' / '_csv'
    output_dir = admin_dir / 'output'
    template_dir = output_dir / 'template'
    
    # 出力フォルダ確認・作成
    template_dir.mkdir(parents=True, exist_ok=True)
    
    # JSON検出
    json_files = list(config_dir.glob('*.json'))
    if not json_files:
        print(f"エラー: config/ にJSONファイルがありません")
        input("\nEnterキーで終了...")
        return
    
    json_path = json_files[0]
    print(f"JSON: {json_path.name}")
    
    # テンプレートファイルパス（_admin/output/template/）
    parsed = parse_json_config(json_path)
    template_name = f"template_{json_path.stem}.xlsx"
    template_path = template_dir / template_name
    
    # 統合ファイルパス（_admin/output/）
    output_name = f"{json_path.stem}_統合.xlsx"
    output_path = output_dir / output_name
    
    # テンプレート生成（既存ならスキップ）
    if template_path.exists():
        print(f"テンプレート: {template_name} (既存のためスキップ)")
    else:
        print(f"テンプレート生成中...")
        total_cols = create_template(parsed, template_path)
        print(f"  -> {template_name} ({total_cols}列)")
    
    # 統合は常にテンプレートから新規作成（リセット方式）
    base_template = template_path
    
    # CSV検出
    if not csv_dir.exists():
        csv_dir.mkdir(exist_ok=True)
        print(f"\ndata/_csv/ フォルダを作成しました")
        print(f"CSVファイルをこのフォルダに配置してください")
        input("\nEnterキーで終了...")
        return
    
    csv_files = list(csv_dir.glob('*.csv'))
    if not csv_files:
        print(f"\ndata/_csv/ にCSVファイルがありません")
        input("\nEnterキーで終了...")
        return
    
    print(f"\nCSV: {len(csv_files)}件")
    for f in csv_files:
        print(f"  - {f.name}")
    
    # 統合実行
    print(f"\n統合処理中...")
    data_count, dup_count = merge_csv_to_template(base_template, csv_files, output_path)
    
    print()
    print("=" * 50)
    print(f"完了: _admin/output/{output_name}")
    print(f"  追加データ: {data_count}件")
    if dup_count:
        print(f"  重複ID: {dup_count}件 (黄色ハイライト)")
    print("=" * 50)
    
    input("\nEnterキーで終了...")


if __name__ == '__main__':
    main()
